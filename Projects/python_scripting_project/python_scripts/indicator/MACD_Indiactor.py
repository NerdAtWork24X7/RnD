
# Import the requests library
import requests
import time
from datetime import datetime
import websocket, json, pprint, talib, numpy
import config
from binance.client import Client
from openpyxl import Workbook
import openpyxl
import os
import telegram_bot
from binance.enums import *

# 15m_50 > 1h_50 Upward trend


# define coin
Coin = "ethusdt"

# Define indicator
indicator_macd = "macd"
indicator_ema = "ema"

SOCKET = "wss://stream.binance.com:9443/ws/" + Coin + "@kline_1m"
connectionstatus = 0
client = Client(config.API_KEY, config.API_SECRET, tld='us')
#websocket.enableTrace(True)

# excel sheet
wbkName = 'ExperimentData.xlsx'

if not os.path.isfile(wbkName):
  wb = Workbook()
  wks = wb.create_sheet(title="TradeHistory")
  wks.cell(row=1, column=1, value="Time")
  wks.cell(row=1, column=2, value="Buy Price")
  wks.cell(row=1, column=4, value="Time")
  wks.cell(row=1, column=5, value="Sell Price")
  wb.save(wbkName)
  wb.close

# Define endpoint
endpoint_Macd = f"https://api.taapi.io/{indicator_macd}"
endpoint_ema = f"https://api.taapi.io/{indicator_ema}"

# Define a parameters dict for the parameters to be sent to the API
parameters_Macd = {
  'secret': config.TAAPI_KEY,
  'exchange': 'binance',
  'symbol': 'ETH/USDT',
  'interval': '1m',
  'optInFastPeriod': '8',
  'optInSlowPeriod': '21',
  'optInSignalPeriod': '5',
}

parameters_ema_1h_50 = {
  'secret': config.TAAPI_KEY,
  'exchange': 'binance',
  'symbol': 'ETH/USDT',
  'interval': '1h',
  'optInTimePeriod': '50',
}

parameters_ema_15m_50 = {
  'secret': config.TAAPI_KEY,
  'exchange': 'binance',
  'symbol': 'ETH/USDT',
  'interval': '15m',
  'optInTimePeriod': '50',
}


def on_open(ws):
  global connectionstatus
  connectionstatus = 1
  print('opened connection')


def on_close(ws):
  global connectionstatus
  connectionstatus = 0
  print('closed connection')


def on_message(ws, message):
  global close_price
  # print('received message')
  json_message = json.loads(message)
  # pprint.pprint(json_message)

  candle = json_message['k']

  is_candle_closed = candle['x']
  close_price = float(candle['c'])
  open_price = float(candle['o'])
  get_stats(open_price, is_candle_closed)


buy_price = 0
sell_price = 0
Buy_Sell_Status = "BUY"
Stop_Loss_Price = 0
Lower_Resistance_Price = 0
Upper_Resistance_Price = 0
Count = 0
trend_ema_15m_50 = 0
trend_ema_1h_50 = 0
enable_flag = 0
Sell_Attempt = 0
MACD_Result_prevCAndle = {}
trend_MACD_buy = -8.0
trend_MACDSignal_buy = -8.0
trend_MACD_sell = 8.0
trend_MACDSignal_sell = 8.0

def Update_Excel(price, Buy_Sell_Status):
  global sell_price,buy_price
  now = datetime.now()
  wbk = openpyxl.load_workbook(wbkName)
  if "TradeHistory" in wbk.sheetnames:
    wks = wbk["TradeHistory"]
  else:
    wks = wbk.create_sheet(title="TradeHistory")

  if Buy_Sell_Status == "BUY":
    for cell in wks["A"]:
      if cell.row is None:
        break
    empty_row = cell.row + 1
    wks.cell(row=empty_row, column=1, value=now.strftime("%H:%M:%S"))
    wks.cell(row=empty_row, column=2, value=price)



  elif Buy_Sell_Status == "SELL":
    for cell in wks["D"]:
      if cell.row is None:
        break
    empty_row = cell.row
    wks.cell(row=empty_row, column=4, value=now.strftime("%H:%M:%S"))
    wks.cell(row=empty_row, column=5, value=price)
    # todo Profit / Loss
    Profit = (sell_price - buy_price) * 0.1
    wks.cell(row=empty_row, column=6, value=str(Profit))
  else:
    pass

  wbk.save(wbkName)
  wbk.close


def execute_Trade(MACD_Result, price, trend):
  global buy_price, sell_price, Buy_Sell_Status, Stop_Loss_Price, Sell_Attempt, MACD_Result_prevCAndle
  global trend_MACD_buy, trend_MACDSignal_buy, trend_MACD_sell, trend_MACDSignal_sell

  if (Buy_Sell_Status == "SELL") and (float(MACD_Result['valueMACD']) > trend_MACD_sell) and (float(MACD_Result['valueMACDSignal']) > trend_MACDSignal_sell):
    if price > buy_price or Sell_Attempt > 60:
      # Check peak
      if bool(MACD_Result_prevCAndle) and (float(MACD_Result['valueMACDHist']) < float(MACD_Result_prevCAndle['valueMACDHist'])):
        sell_price = price
        Sell_Attempt = 0
        Buy_Sell_Status = "BUY"
        # Send Telegram message
        telegram_bot.send_msg_on_telegram("Sold ETH @ " + str(price))
        Update_Excel(price, "SELL")
        print(
          "----------------------------------------------Sell----------------------------------------------------------------")
    else:
      Sell_Attempt += 1
      print(
        "----------------------------------------------Sell Attempt--------------------------------------------------------")

  # check treand and MACD
  elif (trend > 0) and (Buy_Sell_Status == "BUY") and (float(MACD_Result['valueMACD']) < trend_MACD_buy) and (float(MACD_Result['valueMACDSignal']) < trend_MACDSignal_buy):
    # detect steep fall
    if bool(MACD_Result_prevCAndle) and (float(MACD_Result['valueMACDHist']) > float(MACD_Result_prevCAndle['valueMACDHist'])):
      buy_price = price
      Buy_Sell_Status = "SELL"
      telegram_bot.send_msg_on_telegram("Bought ETH @ " + str(price))
      Update_Excel(price, "BUY")
      print(
        "----------------------------------------------Buy------------------------------------------------------------------")

  else:
    pass


def Check_StopLoss_Takeprofit(price):
  global sell_price, buy_price, Buy_Sell_Status
  # TODO
  Stop_Loss_Price = (1 - (8 / 100)) * buy_price
  Profit_Price = (1 + (2 / 100)) * buy_price

  if (Buy_Sell_Status == "SELL"):
    if price < Stop_Loss_Price:
      sell_price = price
      Buy_Sell_Status = "STOP_TRADE"
      Update_Excel(price, "SELL")
      print("Sell")
      telegram_bot.send_msg_on_telegram("Stop loss Triggered ETH @ " + str(price))

    elif price > Profit_Price:
      sell_price = price
      Buy_Sell_Status = "SELL"
      Update_Excel(price, "SELL")
      print("Sell")
      telegram_bot.send_msg_on_telegram("Profit percentage Triggered ETH @ " + str(price))
  '''
  elif price < Lower_Resistance_Price:
    sell_price = price
    Buy_Sell_Status = "STOP_TRADE"
    Update_Excel(price, "SELL")
    print("Sell")

  elif price < Upper_Resistance_Price:
    sell_price = price
    Buy_Sell_Status = "STOP_TRADE"
    Update_Excel(price, "SELL")
    print("Sell")
   '''


prev_Trend = 0

def Send_Trend(trend, price):
  global prev_Trend,trend_MACD_buy, trend_MACDSignal_buy, trend_MACD_sell, trend_MACDSignal_sell
  if prev_Trend < 0 and trend > 0:
    telegram_bot.send_msg_on_telegram("UPward Trend started " + str(price))
  elif prev_Trend > 0 and trend < 0:
    telegram_bot.send_msg_on_telegram("Downward Trend started " + str(price))

  if trend < 20:
    # downward movement
    trend_MACD_buy = -10.0
    trend_MACDSignal_buy = -10.0
    trend_MACD_sell = 10.0
    trend_MACDSignal_sell = 10.0
  elif trend > 20:
    # upward movement
    trend_MACD_buy = -10.0
    trend_MACDSignal_buy = -10.0
    trend_MACD_sell = 10.0
    trend_MACDSignal_sell = 10.0

  prev_Trend = trend


def get_stats(price, is_candle_closed):
  global MACD_Result_prevCAndle,Buy_Sell_Status, Count, trend_ema_15m_50, trend_ema_1h_50, parameters_ema, parameters_Macd, endpoint_ema, endpoint_Macd, enable_flag
  now = datetime.now()

  if enable_flag == 2 and is_candle_closed and int(now.strftime("%S")) < 5:
    # Send get request and save the response as response object
    response = requests.get(url=endpoint_Macd, params=parameters_Macd)
    # Extract data in json format
    MACD_Result = response.json()
    current_time = now.strftime("%H:%M:%S")
    MACD_Data = "  valueMACD = " + str(round(MACD_Result['valueMACD'], 2)) + "  valueMACDSignal = " + str(
      round(MACD_Result['valueMACDSignal'], 2)) + "  valueMACDHist = " + str(round(MACD_Result['valueMACDHist'], 2))
    trend = round((trend_ema_15m_50 - trend_ema_1h_50), 2)
    # Print result
    print(str(current_time) + " price = " + str(price) + MACD_Data + " trend " + str(trend) + "  " + Buy_Sell_Status)

    # send Trend Signal
    Send_Trend(trend, price)
    if int(now.strftime("%M")) == 0:
      telegram_bot.send_msg_on_telegram("Current trend " + str(trend))

    # TODO Exceute Trade
    #execute_Trade(MACD_Result, price, trend)
    execute_Trade(MACD_Result, price, 10)

    # take backup of preious MACD for steepfall
    MACD_Result_prevCAndle = MACD_Result

    # Check Stop Loss and resistance
    Check_StopLoss_Takeprofit(price)

  if int(now.strftime("%S")) == 20:
    # Send get request and save the response as response object
    response = requests.get(url=endpoint_ema, params=parameters_ema_15m_50)
    # Extract data in json format
    result = response.json()
    trend_ema_15m_50 = float(result["value"])
    enable_flag = 1

  if enable_flag == 1 and int(now.strftime("%S")) == 40:
    # Send get request and save the response as response object
    response = requests.get(url=endpoint_ema, params=parameters_ema_1h_50)
    # Extract data in json format
    result = response.json()
    trend_ema_1h_50 = float(result["value"])
    enable_flag = 2
      


while (1):
  if connectionstatus == 0:
    time.sleep(5)
    ws = websocket.WebSocketApp(SOCKET, on_open=on_open, on_close=on_close, on_message=on_message)
    ws.run_forever(ping_interval=70, ping_timeout=10)
    connectionstatus = 0




